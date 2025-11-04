from flask import Flask, render_template, request, jsonify
import smtplib
from email.message import EmailMessage
from openpyxl import load_workbook
import csv
import os

app = Flask(__name__)

UPLOAD_FOLDER = "uploads"
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER

# ✅ Hostinger SMTP Configuration
SMTP_SERVER = "smtp.hostinger.com"
SMTP_PORT = 587  # Use TLS
SENDER_EMAIL = "emma@laskontech.org"
SENDER_PASSWORD = "=VJSv4Ov$"  # same as the working one

@app.route("/")
def index():
    return render_template("index.html")

@app.route("/send_emails", methods=["POST"])
def send_emails():
    try:
        file = request.files["file"]
        email_body = request.form["email_body"]
        file_path = os.path.join(app.config["UPLOAD_FOLDER"], file.filename)
        file.save(file_path)

        print("🔹 Connecting to SMTP (TLS)...")
        with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as smtp:
            smtp.ehlo()
            smtp.starttls()
            smtp.ehlo()
            smtp.login(SENDER_EMAIL, SENDER_PASSWORD)
            print("✅ Logged in successfully.")

            # ✅ CSV or XLSX handling
            if file.filename.endswith(".csv"):
                with open(file_path, newline='', encoding='utf-8') as csvfile:
                    reader = csv.reader(csvfile)
                    next(reader)  # skip header
                    for row in reader:
                        if len(row) < 2:
                            continue
                        name, recipient = row[0].strip(), row[1].strip()
                        body = email_body.replace("{{name}}", name)

                        msg = EmailMessage()
                        msg["From"] = SENDER_EMAIL
                        msg["To"] = recipient
                        msg["Subject"] = "Message from LaskonTech"
                        msg.set_content(body, subtype="html")
                        smtp.send_message(msg)

            elif file.filename.endswith(".xlsx"):
                wb = load_workbook(file_path)
                sheet = wb.active
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if not row or len(row) < 2:
                        continue
                    name, recipient = row
                    body = email_body.replace("{{name}}", str(name))

                    msg = EmailMessage()
                    msg["From"] = SENDER_EMAIL
                    msg["To"] = recipient
                    msg["Subject"] = "Message from LaskonTech"
                    msg.set_content(body, subtype="html")
                    smtp.send_message(msg)
            else:
                return jsonify({
                    "status": "error",
                    "message": "Unsupported file type. Please upload .csv or .xlsx only."
                })

        print("✅ All emails sent successfully.")
        return jsonify({"status": "success"})

    except Exception as e:
        print("❌ Error:", e)
        return jsonify({"status": "error", "message": str(e)})

if __name__ == "__main__":
    app.run(debug=True)
